import os
import pandas as pd
from openpyxl import load_workbook

def set_participant(link, name, phone, num_cheque, date, store, summ):
    
    data = {
        "Телеграмм": [link],
        "Имя": [name],
        "Телефон": [phone],
        "Номер чека": [num_cheque],
        "Дата покупки": [date],
        "Магазин": [store],
        "Сумма покупки": [summ]
    }

    df = pd.DataFrame(data)

    # Используй абсолютный путь
    file_name = r'C:\Users\Никита\Desktop\EnotBot\Participants.xlsx' 
    
    if not os.path.exists(file_name):
        print("Файл не найден, создаем новый файл.")
        try:
            with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            print("Файл успешно создан.")
        except Exception as e:
            print(f"Не удалось создать файл: {e}")
    else:
        try:
            book = load_workbook(file_name)
            visible_sheets = [sheet for sheet in book.sheetnames if book[sheet].sheet_state == 'visible']

            if not visible_sheets:
                raise ValueError("Файл не содержит видимых листов.")

            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = writer.sheets[visible_sheets[0]].max_row
                df.to_excel(writer, index=False, header=False, startrow=startrow)

        except ValueError as ve:
            print(f"Ошибка: {ve}")
        except Exception as e:
            print(f"Произошла ошибка: {e}")

# Пример вызова функции
# set_participant('Телеграм', 'Имя', '1234567890', '1234', '2024-10-12', 'Магазин', 1000)

